import tkinter as tk
from tkinter import ttk, messagebox
import os
import sqlite3 
from datetime import datetime
import openpyxl
from openpyxl.styles import Alignment
import sys

# Class quản lý toàn bộ dữ liệu
class Quan_Ly_Du_Lieu:
    def __init__(self, database):
        self.database = database
    # Phương thức khởi tạo dữ liệu đơn hàng nếu chưa tồn tại, nếu có thì kết nối tới, chạy 1 lần duy nhất khi ứng dụng chạy
    def khoi_tao_du_lieu(self):
        with sqlite3.connect(self.database) as file: # Kết nối với database đặt bí danh là file dùng with tự động đóng kết nối, kể cả khi có lỗi
            con_tro = file.cursor()
            # Tạo bảng đơn hàng nếu chưa tồn tại (IF NOT EXISTS giúp tránh lỗi khi chạy lại)
            # ma_don_hang: Khóa chính- PRIMARY KEY, không được trùng lặp, don_gia: kiểu REAl (số thực) phù hợp cho tiền tệ
            con_tro.execute("""
                CREATE TABLE IF NOT EXISTS don_hang (
                    ma_don_hang TEXT PRIMARY KEY,
                    trang_thai INTEGER,
                    ma_san_pham TEXT,
                    ten_san_pham TEXT,
                    khach_hang TEXT,
                    don_vi_tinh TEXT,
                    don_gia REAL,
                    so_luong INTEGER,
                    ngay_tao TEXT,
                    ngay_hoan_thanh TEXT
                )
            """)
            file.commit()# Lưu lại các thay đổi
            
    # Phương thức lấy toàn bộ đơn hàng từ database, trả về danh sách các tuple, mỗi tuple là 1 đơn hàng
    def tat_ca_don(self):
        with sqlite3.connect(self.database) as file:
            con_tro = file.cursor()
            con_tro.execute("SELECT * FROM don_hang")
            danh_sach = con_tro.fetchall() # lấy ra tất cả dòng kết quả từ câu lệnh truy vấn sql
            return danh_sach # Trả về danh sách để hiển thị

    # Phương thức thêm đơn hàng vào database
    def them_don_hang(self, ma_don_hang, ma_san_pham, ten_san_pham, khach_hang, don_vi_tinh, don_gia, so_luong):
        ngay_tao = datetime.now().strftime("%d/%m/%Y %H:%M") # lấy định dạng ngày/tháng/năm giờ/ phút hiện tại
        with sqlite3.connect(self.database) as file:
            con_tro = file.cursor()
            # INSERT INTO don_hang(...) : khai báo chèn vào các cột được liệt kê trong ngoặc
            # VALUES (?..?) : khai báo 10 dấu ? đại diện giữ chỗ cho 10 cột dữ liệu tương ứng
            # Quá trình xử lý: sqlite bốc các biến trong ngoặc ở dưới truyền vào dấu ?
            # Theo thứ tự từ trái sang phải, biến thứ 1 vào dấu ? số 1,biến thứ 2 vào dấu số 2...
            con_tro.execute("""
                INSERT INTO don_hang (ma_don_hang, trang_thai, ma_san_pham, ten_san_pham, khach_hang, don_vi_tinh, don_gia, so_luong, ngay_tao, ngay_hoan_thanh)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (ma_don_hang, 0, ma_san_pham, ten_san_pham, khach_hang, don_vi_tinh, don_gia, so_luong, ngay_tao, None))
            file.commit()
            
    # Phương thức kiểm tra tồn tại
    def kiem_tra_ton_tai(self, ma_don_hang):
        with sqlite3.connect(self.database) as file:
            con_tro = file.cursor()
            # SELECT * FROM don_hang WHERE ma_don_hang = ?: tìm dòng có mã đơn hàng khớp chính xác 100% so với giá trị truyền vào
            # Dấu ?: placeholder giữ chỗ giúp chống lỗi bảo mật sql injection
            # ma_don_hang, : bắt buộc phải có , để tạo thành bộ dữ liệu tuple 1 phần tử
            con_tro.execute("SELECT * FROM don_hang WHERE ma_don_hang = ?", (ma_don_hang,))
            ket_qua = con_tro.fetchone() # Chỉ lấy ra 1 dòng dữ liệu đầu tiên, tìm thấy trả về dữ liệu nếu có, nếu không trẩ về None
            return ket_qua is not None # Trả về True nếu tìm thấy, False nếu không tìm thấy
    
    # Phương thức sửa đổi thông tin của một đơn hàng đã tồn tại trong database
    def sua_don_hang(self, ma_don_hang, ma_san_pham, ten_san_pham, khach_hang, don_vi_tinh, don_gia, so_luong):
        with sqlite3.connect(self.database) as file:
            con_tro = file.cursor()
            # Thực thi câu lệnh SQL UPDATE để cập nhật dữ liệu, sử dụng các dấu chấm hỏi (?) làm tham số giữ chỗ (placeholder) để tránh lỗi SQL Injection
            con_tro.execute("""
                UPDATE don_hang 
                SET ma_san_pham=?, ten_san_pham=?, khach_hang=?, don_vi_tinh=?, don_gia=?, so_luong=?
                WHERE ma_don_hang=?
            """, (ma_san_pham, ten_san_pham, khach_hang, don_vi_tinh, don_gia, so_luong, ma_don_hang))
            # Lưu ý: Thứ tự các biến trong tuple (bên trên) phải khớp hoàn toàn với thứ tự các dấu (?) trong câu lệnh SQL, ma_don_hang ở cuối dùng để làm tham số xác định mã đơn hàng cần sửa
            file.commit()
        
    # Phương thức xóa đơn hàng khỏi database dựa theo mã đơn hàng
    def xoa_don_hang(self, ma_don_hang):
        with sqlite3.connect(self.database) as file:
            con_tro = file.cursor()
            # DELETE FROM  don_hang: chỉ định xóa dữ liệu trong bảng don_hang
            # WHERE ma_don_hang = ? : điều kiệm bắt buộc để sql chỉ xóa duy nhất dòng có mã bằng ma_don_hang
            con_tro.execute("DELETE FROM don_hang WHERE ma_don_hang = ?", (ma_don_hang,))
            file.commit()
    
    # Phương thức tìm kiếm đơn hàng theo ký tự nhập vào, hàm chỉ truy vấn SELECT (đọc dữ liệu)
    def tim_kiem_don_hang(self, tu_khoa):
        with sqlite3.connect(self.database) as file:
            con_tro = file.cursor()
            # Câu lệnh tìm kiếm tương đối (SELECT LIKE)
            # Ký tự đại diện % (Wildcard): Đại diện cho 1 chuỗi ký tự bất kỳ
            # Cách ghép chuỗi: ('%' + tu_khoa + '%',) : không quan tâm phía trước có gì -  chỉ cần có tu_khoa ở giữa - không quan tâm phía sau có gì
            con_tro.execute("SELECT * FROM don_hang WHERE ma_don_hang LIKE ?", ('%' + tu_khoa + '%',))
            ket_qua = con_tro.fetchall() # lấy ra toàn bộ các dòng kết quả từ câu lệnh truy vấn
            return ket_qua # trả về danh sách các mã đơn hàng có chứa ký tự tu_khoa
    
    # Phương thức tìm kiếm theo khách hàng
    def tim_kiem_khach_hang(self, tu_khoa):
        with sqlite3.connect(self.database) as file:
            con_tro = file.cursor()
            con_tro.execute("SELECT * FROM don_hang WHERE khach_hang LIKE ?", ('%' + tu_khoa + '%',))
            ket_qua = con_tro.fetchall()
            return ket_qua
            
    # Phương thức lấy trạng thái đơn hàng
    def lay_trang_thai(self, ma_don_hang):
        with sqlite3.connect(self.database) as file:
            con_tro = file.cursor()
            con_tro.execute("SELECT trang_thai FROM don_hang WHERE ma_don_hang = ?", (ma_don_hang,))
            # fetchone(): lấy ra dòng kết quả đầu tiên tìm thấy, trả về tuple chứa 1 phần tử nếu có, trả về None nếu không
            ket_qua = con_tro.fetchone()
            if ket_qua is not None:
                return ket_qua[0] # lấy phần tử đầu tiên (index 0) trong tuple, chính là giá trị trạng thái
            else:
                return None # trả về none nếu không tìm thấy
            
    # Phương thức cập nhật trạng thái đơn hàng
    def cap_nhat_trang_thai(self, ma_don_hang, trang_thai_moi):
        # Nếu chuyển sang hoàn thành (1): lưu thời gian hoàn thành
        # Nếu bỏ hoàn thành (0): xóa thời gian hoàn thành về None
        if trang_thai_moi == 1:
            ngay_hoan_thanh = datetime.now().strftime("%d/%m/%Y %H:%M")
        else:
            ngay_hoan_thanh = None
        with sqlite3.connect(self.database) as file:
            con_tro = file.cursor()
            # UPDATE don_hang SET trang_thai = ?: cập nhật giá trị mới cho cột trang_thai
            # WHERE ma_don_hang= ? Điều kiện bắt buộc để chỉ áp dụng thay đổi trên 1 đơn hàng duy nhất có mã = mã đơn hàng nhập vào
            con_tro.execute("UPDATE don_hang SET trang_thai = ?, ngay_hoan_thanh = ? WHERE ma_don_hang = ?", (trang_thai_moi, ngay_hoan_thanh, ma_don_hang))
            file.commit()
            
# Class đăng nhập trước khi thao tác chương trình chính
class Cua_So_Dang_Nhap:
    # Phương thức khởi tạo đối tượng ung_dung chính làm tham số để gọi lại sau khi đăng nhập thành công
    def __init__(self, ung_dung):
        self.ung_dung = ung_dung
        # tạo cửa số đăng nhập popup bằng toplevel
        self.cua_so = tk.Toplevel()
        self.cua_so.title("Đăng nhập")
        self.cua_so.minsize(440,220)
        self.cua_so.transient(self.ung_dung.giao_dien) # Khai báo đây là cửa sổ con phụ thuộc vào giao_dien
        # Căn giữa với cửa sổ chính
        self.ung_dung.giao_dien.update_idletasks() # Đảm bảo tính toán kích thước thật của giao_dien, tránh lỗi khi kéo dãn
        x = self.ung_dung.giao_dien.winfo_x() + (self.ung_dung.giao_dien.winfo_width() // 2) - 220 # (lấy tọa độ bên trái + 1 nửa chiều dài của giao_dien) - 1 nửa chiều dài đăng nhập
        y = self.ung_dung.giao_dien.winfo_y() + (self.ung_dung.giao_dien.winfo_height() // 2) -110 # tương tự như trên nhưng là chiều cao
        self.cua_so.geometry(f"440x220+{x}+{y}") # Thiết lập kích thước (rộng 440px, cao 220) và vị trí xuất hiện dựa trên tọa độ x, y vừa tính
        # gọi phương thức xây dựng giao diện
        self.tao_giao_dien()

    # Phương thức xây dựng toàn bộ giao diện cửa sổ đăng nhập
    def tao_giao_dien(self):
        menu_dang_nhap = tk.Frame(self.cua_so)
        menu_dang_nhap.pack(pady=10)
        tk.Label(menu_dang_nhap, text="Tài khoản:", font=("Arial", 10)).grid(row=0, column=0, pady=(10, 5), padx=(10, 10))
        self.nhap_tai_khoan = ttk.Entry(menu_dang_nhap, width=15, style="nhapdulieu.TEntry")
        self.nhap_tai_khoan.grid(row=0, column=1, pady=(10, 10), padx=(5, 10))
        tk.Label(menu_dang_nhap, text="Mật khẩu:", font=("Arial", 10)).grid(row=1, column=0, pady=(5), padx=(10, 10))
        self.nhap_mat_khau = ttk.Entry(menu_dang_nhap, width=15, style="nhapdulieu.TEntry", show="*")
        self.nhap_mat_khau.grid(row=1, column=1, pady=(5,10), padx=(5, 10))
        self.nhan_thong_bao = tk.Label(menu_dang_nhap, text="Nhập thông tin tài khoản\nTài khoản: admin | pass: 123", fg="#555555", font=("Arial", 10))
        self.nhan_thong_bao.grid(row=3, column=0, columnspan=2, pady=(10))
        nut_dang_nhap = tk.Button(menu_dang_nhap, text="Đăng nhập", width=25, bg="lightgreen", fg="#000000", font=("Arial", 10, "bold"), relief="raised", bd=2, command=self.xu_ly_dang_nhap)
        nut_dang_nhap.grid(row=2, column=0, columnspan=2, pady=5, ipady=1)
        # bind(): phương thức đăng ký sự kiện cho widget
        # '<Return>': tên sự kiện - Cho phép nhấn Enter để đăng nhập
        # lambda event: hàm vô danh, nhận tham số event ( bắt buộc để hoạt động )
        # self.xu_ly_dang_nhap(): hàm xử lý logic đăng nhập
        self.cua_so.bind('<Return>', lambda event: self.xu_ly_dang_nhap())
        
    # Phương thức xử lý đăng nhập
    def xu_ly_dang_nhap(self):
        tai_khoan = self.nhap_tai_khoan.get().strip().upper()
        mat_khau = self.nhap_mat_khau.get().strip()

        if tai_khoan == "ADMIN" and mat_khau == "123":
            self.nhan_thong_bao.config(text="Đăng nhập thành công", fg="green")
            self.cua_so.after(500, self.cua_so.destroy)
            self.ung_dung.dang_nhap_thanh_cong()
        elif tai_khoan =="" or mat_khau =="":
            self.nhan_thong_bao.config(text="Vui lòng nhập đầy đủ thông tin\nTài khoản: admin | pass: 123", fg="red")
        else:
            self.nhan_thong_bao.config(text="Thông tin đăng nhập chưa chính xác\nTài khoản: admin | pass: 123", fg="red")
            self.nhap_mat_khau.delete(0, tk.END)
      
# Class chính chứa toàn bộ giao diện, xử lý logic và chức năng
class Quan_Ly_Don_Hang:
    # Phương thức khởi tạo cửa sổ chính, xác định đường dẫn file, tạo đối tượng database và xây dựng giao diện
    def __init__(self):
        # tạo và cấu hình cửa sổ giao diện chính
        self.giao_dien = tk.Tk()
        self.giao_dien.title("Quản lý đơn hàng sản xuất")
        self.giao_dien.minsize(1100,650)
        
        # tạo style dùng chung cho các ô nhập dữ liệu, và bảng hiển thị treeview
        self.style = ttk.Style()
        # Cấu hình riêng cho ô nhập liệu Entry,padding giúp chữ không bị dính sát vào viền ô, dấu * là đại diện cho widget TEntry, với tất cả các thuộc tính liên quan đến Font
        self.style.configure("nhapdulieu.TEntry", padding=(5, 3, 5, 3))
        self.giao_dien.option_add("*TEntry*Font", "Arial 9")
        # Cấu hình chiều cao hàng dữ liệu và ép font cho Treeview
        self.style.configure("Treeview", rowheight=25, font=("Arial", 10))
        self.style.configure("Treeview.Heading", font=("Arial", 10))
        
        # Lấy đường dẫn file dữ liệu
        if getattr(sys, 'frozen', False): # Nếu chạy bằng file exe: kiểm tra xem chương trình chạy ở môi trường nào
            thu_muc_goc = os.path.dirname(sys.executable) # sys.excutable trả về đường dẫn đầy đủ của chính file .exe, os.path.dirname(...) lấy ra thư mục chứa file exe làm thư mục gỗc
        else:
            # __file__ là biến hệ thống chứa đường dẫn file py hiện tại
            # os.path.abspath(__file__): lấy đường dẫn tuyệt đối
            thu_muc_goc = os.path.dirname(os.path.abspath(__file__))
        # Tự động nối thu_muc_goc với file database để tạo đường dẫn
        # Việc dùng os.path.join giúp code chạy được trên cả windows ( dùng dấu \ ) lẫn macos và linux ( dùng dấu /) mà không gây lỗi
        self.file_du_lieu = os.path.join(thu_muc_goc, "du_lieu_don_hang.db")
        self.file_bao_cao = os.path.join(thu_muc_goc, "bao_cao_don_hang.xlsx")
                    
        # Tạo đối tượng quản lý dữ liệu,tất cả thao tác sql sẽ thông qua đối tượng này
        self.quan_ly_du_lieu = Quan_Ly_Du_Lieu(self.file_du_lieu)

        # Gọi phương thức xây dựng giao diện
        self.tao_giao_dien()

    # Phương thức xây dựng toàn bộ giao diện chính
    def tao_giao_dien(self):
        # Tạo menu chứa các ô nhập dữ liệu
        menu_nhap = tk.Frame(self.giao_dien)
        menu_nhap.pack(pady=(15,5))
        # Dòng nhập liệu 1:
        tk.Label(menu_nhap, text="Mã sản phẩm:", font=("Arial", 10)).grid(row=0, column=0, padx=(0,5), pady=(5))
        self.nhap_ma_san_pham = ttk.Entry(menu_nhap, width=14, style="nhapdulieu.TEntry")
        self.nhap_ma_san_pham.grid(row=0, column=1, padx=(0,5))

        tk.Label(menu_nhap, text="Tên sản phẩm:", font=("Arial", 10)).grid(row=0, column=2, padx=(10,5), pady=(5))
        self.nhap_ten_san_pham = ttk.Entry(menu_nhap, width=14, style="nhapdulieu.TEntry")
        self.nhap_ten_san_pham.grid(row=0, column=3, padx=(0,5))

        tk.Label(menu_nhap, text="Khách hàng:", font=("Arial", 10)).grid(row=0, column=4, padx=(10,5), pady=(5))
        self.nhap_khach_hang = ttk.Entry(menu_nhap, width=14, style="nhapdulieu.TEntry")
        self.nhap_khach_hang.grid(row=0,column=5)
        
        # Dòng nhập liệu 2
        tk.Label(menu_nhap, text="Đơn vị tính:", font=("Arial", 10)).grid(row=1, column=0, padx=(0, 5))
        self.nhap_don_vi_tinh = ttk.Entry(menu_nhap, width=14, style="nhapdulieu.TEntry")
        self.nhap_don_vi_tinh.grid(row=1, column=1, padx=(0, 5), pady=(8,0))

        tk.Label(menu_nhap, text="Số lượng:", font=("Arial", 10)).grid(row=1, column=2, padx=(0, 5))
        self.nhap_so_luong = ttk.Entry(menu_nhap, width=14, style="nhapdulieu.TEntry")
        self.nhap_so_luong.grid(row=1, column=3, padx=(0, 5), pady=(8,0))

        tk.Label(menu_nhap, text="Đơn giá (VNĐ):", font=("Arial", 10)).grid(row=1, column=4, padx=(10, 5), pady=(5))
        self.nhap_don_gia = ttk.Entry(menu_nhap, width=14, style="nhapdulieu.TEntry")
        self.nhap_don_gia.grid(row=1, column=5, pady=(8,0))
        
        # Tạo menu tiện ích: tìm kiếm, hoàn thành, hiển thị, xuất báo cáo
        menu_tien_ich = tk.Frame(self.giao_dien)
        menu_tien_ich.pack(pady=(10,5))

        tk.Label(menu_tien_ich, text="Mã đơn hàng / tìm kiếm:", font=("Arial", 10)).pack(side="left", padx=(0, 5))
        self.nhap_ma_don_hang = ttk.Entry(menu_tien_ich, width=14, style="nhapdulieu.TEntry")
        self.nhap_ma_don_hang.pack(side="left", padx=(10), pady=(1,0))
        
        self.nut_tim_kiem = tk.Button(menu_tien_ich, text="🔍 Tìm kiếm ▾", width=16, bg="#d5e1f9", fg="#000000", font=("Arial", 9, "bold"), relief="raised", bd=2)
        self.nut_tim_kiem.pack(side="left", padx=10)

        self.nut_hoan_thanh = tk.Button(menu_tien_ich, text="✅ Hoàn thành", width=16, bg="#c2fbd7", fg="#000000", font=("Arial", 9, "bold"), relief="raised", bd=2)
        self.nut_hoan_thanh.pack(side="left", padx=(10))

        self.nut_hien_thi = tk.Button(menu_tien_ich, text="📋 Tất cả đơn", width=16, bg="#d5e1f9", fg="#000000", font=("Arial", 9, "bold"), relief="raised", bd=2)
        self.nut_hien_thi.pack(side="left", padx=10)
        
        self.nut_xuat_bao_cao = tk.Button(menu_tien_ich, text="⬇️ Xuất báo cáo", width=16, bg="#9ff0d9", fg="#000000", font=("Arial", 9, "bold"), relief="raised", bd=2)
        self.nut_xuat_bao_cao.pack(side="left", padx=(10, 0))

        # Tạo menu chức năng: thêm, sửa, xóa đơn hàng
        menu_chuc_nang = tk.Frame(self.giao_dien)
        menu_chuc_nang.pack(pady=10)
        
        self.nut_them_don = tk.Button(menu_chuc_nang, text="➕Thêm đơn hàng", width=20, bg="#9ff0d9", fg="#000000", font=("Arial", 9, "bold"), relief="raised", bd=2)
        self.nut_them_don.pack(side="left", padx=(0, 10))
        
        self.nut_sua_don = tk.Button(menu_chuc_nang, text="✏️ Sửa đơn hàng", width=20, bg="#f9e79f", fg="#000000", font=("Arial", 9, "bold"), relief="raised", bd=2)
        self.nut_sua_don.pack(side="left", padx=(10, 10))

        self.nut_xoa_don = tk.Button(menu_chuc_nang, text="🗑️ Xóa đơn hàng", width=20, bg="#EF4444", fg="#F3F0F0", font=("Arial", 9, "bold"), relief="raised", bd=2)
        self.nut_xoa_don.pack(side="left", padx=(10, 0))
        
        # Hiển thị tổng đơn, kết quả tìm kiếm
        self.nhan_tu_khoa = tk.Label(self.giao_dien, text="Quản lý đơn hàng", font=("Arial", 11, "bold"), fg="green")
        self.nhan_tu_khoa.pack(pady=(0))
        
        self.tong_don = tk.Label(self.giao_dien, text="Tổng đơn hàng: 0", font=("Arial", 11, "bold"), fg="green")
        self.tong_don.pack(pady=(0))

        # Tạo bảng hiển thị dữ liệu Treeview
        self.hien_thi_du_lieu = ttk.Treeview(
            self.giao_dien,
            columns=("Đơn hàng", "Trạng thái", "Mã sản phẩm", "Sản phẩm", "Khách hàng", "Đơn vị tính", "Đơn giá (VNĐ)", "Số lượng", "Thành tiền (VNĐ)", "Ngày tạo đơn", "Ngày hoàn thành"),
            show="headings", # Chỉ hiển thị tiêu đề và dữ liệu, ẩn cột ID/Index mặc đinhj
            height=15        # Chiều cao cố định 15 dòng dữ liệu, nhiều hơn sẽ tự động ẩn để cuộn
        )
        # Tạo tiêu đề và set độ rộng cho cột
        for ten_cot in self.hien_thi_du_lieu["columns"]:
            self.hien_thi_du_lieu.heading(ten_cot, text=ten_cot) # tên cột, tiêu đề dùng tên cột tại columns để tạo
            self.hien_thi_du_lieu.column(ten_cot, width=100)
        # Tùy chỉnh độ rộng 1 số cột đặc biệt
        self.hien_thi_du_lieu.column("Đơn hàng", width=80)
        self.hien_thi_du_lieu.column("Trạng thái", width=120)
        self.hien_thi_du_lieu.column("Đơn vị tính", width=80)
        self.hien_thi_du_lieu.column("Số lượng", width=80)
        self.hien_thi_du_lieu.column("Thành tiền (VNĐ)", width=120)
        self.hien_thi_du_lieu.column("Ngày tạo đơn", width=120)
        self.hien_thi_du_lieu.column("Ngày hoàn thành", width=120)
        # Đưa bảng dữ liệu lên giao diện chính
        # fill = "both" : kéo giãn bảng theo 2 chiều để lấp đầy không gian trống theo giao_dien, 
        # Expand = True: cho phép bảng mở rộng thêm kích thước khi người dùng phóng to thu nhỏ
        self.hien_thi_du_lieu.pack(padx=10, pady=10, fill="both", expand=True)
    
    # Phương thức cập nhật dữ liệu, hiển thị lên giao diện
    def cap_nhat_du_lieu(self, du_lieu_hien_thi = None):
        # Xóa dữ liệu cũ đang có trên treeview để tránh bị trùng lặp
        for hang in self.hien_thi_du_lieu.get_children(): # Lấy ra danh sách ID của tất cả các dòng hiện có trên treview
            self.hien_thi_du_lieu.delete(hang)
        # Lấy dữ liệu hiển thị, dùng danh sách truyền vào nếu có, nếu không thì lấy toàn bộ từ database
        if du_lieu_hien_thi is not None:
            danh_sach_hien_thi = du_lieu_hien_thi# sử dụng danh sách truyền vào, chức năng tìm kiém
        else:
            danh_sach_hien_thi = self.quan_ly_du_lieu.tat_ca_don() # gọi phương thức lấy toàn bộ dữ liệu từ database
        # Duyệt qua từng đơn và tách các cột dữ liệu theo vị trí index
        for hang in danh_sach_hien_thi:
            ma_don_hang = hang[0]
            trang_thai_so = hang[1]
            ma_san_pham = hang[2]
            ten_san_pham = hang[3]
            khach_hang = hang[4]
            don_vi_tinh = hang[5]
            don_gia = hang[6]
            so_luong = hang[7]
            ngay_tao = hang[8]
            if hang[9] is not None and hang[9] != "":
                # Nếu hang[9] không phải None và không phải chuỗi rỗng
                ngay_hoan_thanh = hang[9]
            else:
                # Nếu hang[9] là None hoặc là chuỗi rỗng
                ngay_hoan_thanh = "---"
            
            if trang_thai_so==1:
                trang_thai_chu = "✅Hoàn thành"
            else:
                trang_thai_chu = "⌛Chưa hoàn thành"
            
            thanh_tien = don_gia * so_luong

            # Chèn dòng dữ liệu mới vào cuối bảng treeview, "": chèn ở cấp cao nhất, mọi đơn hàng cùng cấp với nhau, end: chèn từ dưới lên
            self.hien_thi_du_lieu.insert("", "end", values=(
                ma_don_hang,
                trang_thai_chu,
                ma_san_pham,
                ten_san_pham,
                khach_hang,
                don_vi_tinh,
                don_gia,
                so_luong,
                thanh_tien,
                ngay_tao,
                ngay_hoan_thanh
            ))
        # Lấy số liệu lên label tổng đơn hàng
        so_don = len(self.hien_thi_du_lieu.get_children())
        self.tong_don.config(text=f"Tổng đơn hàng: {so_don}")
    
    # Phương thức xóa dữ liệu nhập tại các ô nhập liệu, xóa từ vị trí index 0  cho đến end
    def xoa_du_lieu_nhap(self):
        self.nhap_ma_san_pham.delete(0, tk.END)
        self.nhap_ten_san_pham.delete(0, tk.END)
        self.nhap_khach_hang.delete(0, tk.END)
        self.nhap_don_vi_tinh.delete(0, tk.END)
        self.nhap_don_gia.delete(0, tk.END)
        self.nhap_so_luong.delete(0, tk.END)
        self.nhap_ma_don_hang.delete(0, tk.END)

    # Phương thức thêm đơn hàng mới vào database
    def them_don_hang(self):
        ma_don_hang = self.nhap_ma_don_hang.get().strip().upper() # xóa khoảng trắng 2 đầu, viết hoa toàn bộ ký tự
        ma_san_pham = self.nhap_ma_san_pham.get().strip().upper()
        ten_san_pham = self.nhap_ten_san_pham.get().strip().title() # xóa khoảng trắng 2 đầu, viết hoa ký từ đầu mỗi từ
        khach_hang = self.nhap_khach_hang.get().strip().upper()
        don_vi_tinh = self.nhap_don_vi_tinh.get().strip().title()
        so_luong = self.nhap_so_luong.get().strip()
        don_gia = self.nhap_don_gia.get().strip()
        
        # Kiểm tra ô nhập liệu không được bỏ trống
        if not all([ma_don_hang, ma_san_pham, ten_san_pham, khach_hang, don_vi_tinh,don_gia, so_luong]):
            messagebox.showwarning("Thiếu thông tin", "Vui lòng nhập đầy đủ thông tin")
            return
        try: # kiểm tra định dạng số với đơn giá và số lượng
            don_gia = float(don_gia)
            so_luong = int(so_luong)
            if don_gia <= 0 or so_luong <= 0:
                messagebox.showwarning("Lỗi", "Giá và số lượng phải lớn hơn 0")
                return
        except ValueError:
            messagebox.showwarning("Lỗi", "Giá phải là số\nSố lượng phải là số nguyên")
            return
        
        # Kiểm tra đơn hàng tồn tại chưa bằng cách gọi hàm từ database
        if self.quan_ly_du_lieu.kiem_tra_ton_tai(ma_don_hang):
            messagebox.showwarning("Lỗi", "Đã tồn tại đơn hàng này\nVui lòng tạo đơn hàng khác")
            return
        # Gọi phương thức them_don_hang vào database qua đối tượng quan_ly_du_lieu
        self.quan_ly_du_lieu.them_don_hang(ma_don_hang, ma_san_pham, ten_san_pham, khach_hang, don_vi_tinh, don_gia,so_luong)

        # làm mới lại bảng hiển thị và xóa dữ liệu nhập tại các ô
        self.cap_nhat_du_lieu()
        self.xoa_du_lieu_nhap()
        messagebox.showinfo("Thành công", f"Đã tạo đơn hàng mới: {ma_don_hang} với sản phẩm {ten_san_pham}")
        
    # Phương thức sửa , cập nhật thông tin đơn hàng
    def sua_don_hang(self):
        ma_don_hang = self.nhap_ma_don_hang.get().strip().upper()
        ma_san_pham = self.nhap_ma_san_pham.get().strip().upper()
        ten_san_pham = self.nhap_ten_san_pham.get().strip().title()
        khach_hang = self.nhap_khach_hang.get().strip().upper()
        don_vi_tinh = self.nhap_don_vi_tinh.get().strip().title()
        so_luong = self.nhap_so_luong.get().strip()
        don_gia = self.nhap_don_gia.get().strip()

        # Kiểm tra đầu vào
        if not all([ma_don_hang, ma_san_pham, ten_san_pham, khach_hang, don_vi_tinh, don_gia, so_luong]):
            messagebox.showwarning("Thiếu thông tin", "Vui lòng nhập đầy đủ thông tin")
            return
        if not self.quan_ly_du_lieu.kiem_tra_ton_tai(ma_don_hang):
            messagebox.showwarning("Lỗi", f"Không tìm thấy đơn hàng: {ma_don_hang}")
            return
        try:
            don_gia = float(don_gia)
            so_luong = int(so_luong)
            if don_gia <= 0 or so_luong <= 0:
                messagebox.showwarning("Lỗi", "Giá và số lượng phải lớn hơn 0")
                return
        except ValueError:
            messagebox.showwarning("Lỗi", "Giá phải là số\nSố lượng phải là số nguyên")
            return

        # Xác nhận và thực hiện sửa
        xac_nhan_sua = messagebox.askyesno("Xác nhận", f"Xác nhận sửa đơn hàng {ma_don_hang}?")
        if xac_nhan_sua == True:
            self.quan_ly_du_lieu.sua_don_hang(ma_don_hang, ma_san_pham, ten_san_pham, khach_hang, don_vi_tinh, don_gia, so_luong)
            self.cap_nhat_du_lieu()
            self.xoa_du_lieu_nhap()
            messagebox.showinfo("Thành công", f"Đã cập nhật đơn hàng {ma_don_hang}")
        else:
            messagebox.showinfo("Hủy", f"Bạn đã hủy thao tác sửa đơn hàng {ma_don_hang}")

    # Phương thức xóa đơn hàng khỏi database
    def xoa_don_hang(self):
        tu_khoa = self.nhap_ma_don_hang.get().strip().upper()
        if tu_khoa == "":
            messagebox.showwarning("Lỗi", "Vui lòng nhập mã đơn hàng muốn xóa")
            self.xoa_du_lieu_nhap()
            return
        if not self.quan_ly_du_lieu.kiem_tra_ton_tai(tu_khoa):
            messagebox.showwarning("Lỗi", f"Không tìm thấy đơn hàng nào có mã là: {tu_khoa}")
            self.xoa_du_lieu_nhap()
            return
        # Hỏi xác nhận trước khi xóa
        xac_nhan_xoa = messagebox.askyesno("Xác nhận xóa", f"Bạn có chắc muốn xóa đơn hàng {tu_khoa} này không?")
        if xac_nhan_xoa == True:
            self.quan_ly_du_lieu.xoa_don_hang(tu_khoa)
            self.cap_nhat_du_lieu()
            self.xoa_du_lieu_nhap()
            messagebox.showinfo("Thành công", f"Đã xóa thành công đơn hàng: {tu_khoa}")
        else:
            messagebox.showinfo("Hủy", f"Bạn đã hủy thao tác xóa đơn hàng {tu_khoa}")

    # Phương thức tìm kiếm đơn hàng
    def tim_kiem_don_hang(self, cach_tim="ma_don_hang"):
        tu_khoa = self.nhap_ma_don_hang.get().strip().upper()
        if tu_khoa == "":
            messagebox.showwarning("Thiếu thông tin", "Vui lòng nhập từ khóa vào ô tìm kiếm")
            return

        if cach_tim == "ma_don_hang":
            # Tìm theo mã đơn hàng
            danh_sach_tim_kiem = self.quan_ly_du_lieu.tim_kiem_don_hang(tu_khoa)
            self.nhan_tu_khoa.config(text=f"Kết quả tìm kiếm: {tu_khoa}")
            thong_bao = f"mã đơn hàng chứa '{tu_khoa}'"
        else:
            # Nếu không tìm theo mã đơn hàng thì sẽ tìm theo khách hàng
            danh_sach_tim_kiem = self.quan_ly_du_lieu.tim_kiem_khach_hang(tu_khoa)
            self.nhan_tu_khoa.config(text=f"Khách hàng: {tu_khoa}")
            thong_bao = f"khách hàng chứa '{tu_khoa}'"

        if len(danh_sach_tim_kiem) > 0:
            self.cap_nhat_du_lieu(danh_sach_tim_kiem)
            messagebox.showinfo("Kết quả", f"Tìm thấy {len(danh_sach_tim_kiem)} đơn hàng có {thong_bao}")
            self.xoa_du_lieu_nhap()
        else:
            messagebox.showwarning("Kết quả", f"Không tìm thấy đơn hàng nào có {thong_bao}")
            self.nhan_tu_khoa.config(text="")
            self.cap_nhat_du_lieu()
            self.xoa_du_lieu_nhap()
            
    # Phương thức xử lý nút tìm kiếm
    def xu_ly_tim_kiem(self):
        menu_tim_kiem = tk.Menu(self.giao_dien, tearoff=0) # Tạo menu con khi tìm kiếm, tearoff=0 ép menu con dính sát nút tìm kiếm
        menu_tim_kiem.add_command(label="Tìm theo mã đơn hàng", command=lambda: self.tim_kiem_don_hang("ma_don_hang")) # hàm ấn danh lambda truyền tham số ( tìm theo mã đơn hàng/ khách hàng ) cho command
        menu_tim_kiem.add_command(label="Tìm theo khách hàng",  command=lambda: self.tim_kiem_don_hang("khach_hang"))
        # Gán chức năng hành động cho nút tìm kiếm, menu_tim_kiem.post() hiển thị menu tìm kiếm tại vị trí x,y trên màn hình
        self.nut_tim_kiem.config(command=lambda: menu_tim_kiem.post(
            self.nut_tim_kiem.winfo_rootx(), # lấy tọa độ x (viền bên trái nút tìm kiếm)
            self.nut_tim_kiem.winfo_rooty() + self.nut_tim_kiem.winfo_height() # lấy tọa độ y viền bên trên + lấy chiều cao nút tìm kiếm
        ))

    # Phương thức hiển thị tất cả
    def hien_thi_tat_ca(self):
        self.nhan_tu_khoa.config(text="Quản lý đơn hàng")
        self.xoa_du_lieu_nhap()
        self.cap_nhat_du_lieu()
    
    # Phương thức thay đổi trạng thái đơn hàng : hoàn thành / không hoàn thành
    def hoan_thanh_don(self):
        tu_khoa = self.nhap_ma_don_hang.get().strip().upper()
        if tu_khoa == "":
            messagebox.showwarning("Lỗi", "Vui lòng nhập mã đơn hàng cần xác nhận hoàn thành")
            self.xoa_du_lieu_nhap()
            return
        # Gọi phương thức lấy trạng thái hiện tại từ đối tượng database
        trang_thai_hien_tai = self.quan_ly_du_lieu.lay_trang_thai(tu_khoa)
        if trang_thai_hien_tai is None:
            # Không tìm thấy đơn hàng
            messagebox.showwarning("Lỗi", "Mã đơn hàng không đúng\nHãy nhập đúng mã đơn hàng")
            self.xoa_du_lieu_nhap()
            return
        if trang_thai_hien_tai == 1:
            trang_thai_moi = 0
            trang_thai_chu = "Chưa hoàn thành"
        else:
            trang_thai_moi = 1
            trang_thai_chu = "Hoàn thành"
        # Gọi phương thức cập nhật trạng thái từ đối tượng database
        self.quan_ly_du_lieu.cap_nhat_trang_thai(tu_khoa, trang_thai_moi)
        self.cap_nhat_du_lieu()
        messagebox.showinfo("Thành công", f"Đơn hàng {tu_khoa} đã chuyển sang trạng thái {trang_thai_chu}")
        self.xoa_du_lieu_nhap()

    # Phương thức xuất báo cáo
    def xuat_bao_cao(self):
        # Gọi phương thức lấy toàn bộ đơn hàng từ đối tượng database
        danh_sach_don_hang = self.quan_ly_du_lieu.tat_ca_don()
        if len(danh_sach_don_hang) == 0:
            messagebox.showwarning("Thông báo", "Không có dữ liệu để xuất báo cáo")
            return

        try:
            wb = openpyxl.Workbook() # Tạo file excel mới trong ram
            sheet = wb.active # chọn sheet đầu tiên
            sheet.title = "Báo cáo đơn hàng"
            
            # Tạo style căn lề để sử dụng cho các dòng, cột trong excel
            can_giua = Alignment(horizontal="center", vertical="center")
            can_trai = Alignment(horizontal="left", vertical="center")
            can_phai = Alignment(horizontal="right", vertical="center")
            
            # Ghi tiêu đề các cột
            tieu_de_cot = ["Đơn hàng", "Trạng thái", "Mã sản phẩm", "Sản phẩm", "Khách hàng", "Đơn vị tính", "Đơn giá (VNĐ)", "Số lượng", "Thành tiền (VNĐ)", "Ngày tạo đơn", "Ngày hoàn thành"]
            # Duyệt qua từng tiêu đề và ghi vào dòng đầu tiên của excel, enumerate(...1) giúp lấy ra số cột, bắt đầu từ (A=1, B=2, C=3, ..)
            for so_cot, tieu_de in enumerate(tieu_de_cot, 1):
                sheet.cell(row=1, column=so_cot, value=tieu_de).alignment = can_giua # ghi tiêu đê và ô tương ứng ở dòng 1 đông thời căn giữa
            
            # Ghi dữ liệu từng đơn hàng: duyệt qua từng đơn hàng trong danh sách, dữ liệu ghi từ dòng số 2 ( start=2), enumerate: tạo bộ đếm tự động
            for chi_so_dong, hang in enumerate(danh_sach_don_hang, start=2):
                if hang[1] ==1:
                    trang_thai_chu = "✅ Hoàn thành"
                else:
                    trang_thai_chu = "Chưa hoàn thành"
                thanh_tien = hang[6] * hang[7]

                # Ghi dữ liệu từng ô cụ thể trên dòng hiện tại (chi_so_dong) và thiết lập căn lề, định dạng số
                sheet.cell(row=chi_so_dong, column=1, value=hang[0]).alignment = can_giua # mã đơn hàng
                sheet.cell(row=chi_so_dong, column=2, value=trang_thai_chu).alignment = can_giua # trạng thái
                sheet.cell(row=chi_so_dong, column=3, value=hang[2]).alignment = can_trai # mã sản phẩm
                sheet.cell(row=chi_so_dong, column=4, value=hang[3]).alignment = can_trai # Tên sản phẩm
                sheet.cell(row=chi_so_dong, column=5, value=hang[4]).alignment = can_trai # Khách hàng
                sheet.cell(row=chi_so_dong, column=6, value=hang[5]).alignment = can_giua # Đơn vị tính
                cell_don_gia = sheet.cell(row=chi_so_dong, column=7, value=hang[6]) # Đơn giá
                cell_don_gia.alignment = can_phai
                cell_don_gia.number_format = '#,##0' # định dạng số cho đơn giá (vd: 2000=2,000)
                cell_so_luong = sheet.cell(row=chi_so_dong, column=8, value=hang[7]) # Số lượng
                cell_so_luong.alignment = can_giua
                cell_so_luong.number_format = '#,##0'
                cell_thanh_tien = sheet.cell(row=chi_so_dong, column=9, value=thanh_tien) # Thành tiền 
                cell_thanh_tien.alignment = can_phai
                cell_thanh_tien.number_format = '#,##0'
                sheet.cell(row=chi_so_dong, column=10, value=hang[8]).alignment = can_giua # ngày tạo đơn
                sheet.cell(row=chi_so_dong, column=11, value=hang[9]).alignment = can_giua # ngày hoàn thành
            
            #Tính tổng số lượng và tổng thành tiền, duyệt lại danh sách đơn hàng để cộng dồn giá trị
            tong_so_luong = 0
            tong_thanh_tien = 0
            for hang in danh_sach_don_hang:
                tong_so_luong += hang[7]
                tong_thanh_tien += (hang[6] * hang[7]) # don_gia * so_luong
                
            # xác định vị trí dòng ghi tổng ( số đơn hàng + 1 dòng tiêu đề + 1 dòng cách ra + 1 dòng chính nó = len()+ 3)
            dong_tong = len(danh_sach_don_hang) + 3
            sheet.cell(row=dong_tong, column=6, value="Tổng:").alignment = can_trai
            cell_tong_don = sheet.cell(row=dong_tong, column=7, value=f"{len(danh_sach_don_hang)} đơn hàng")
            cell_tong_don.alignment = can_phai
            cell_tong_don.number_format =  '#,##0'
            cell_tong_so_luong = sheet.cell(row=dong_tong, column=8, value=tong_so_luong)
            cell_tong_so_luong.alignment = can_giua
            cell_tong_so_luong.number_format = '#,##0'
            cell_tong_thanh_tien = sheet.cell(row=dong_tong, column=9, value=tong_thanh_tien)
            cell_tong_thanh_tien.alignment = can_phai
            cell_tong_thanh_tien.number_format = '#,##0'
            
            # duyệt qua các cột từ A-I để chỉnh lại độ rộng cột trong excel
            for ky_tu_cot in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']:
                sheet.column_dimensions[ky_tu_cot].width = 16 # đặt độ dài bằng 16 tránh chữ bị che khuất
            
            # Lưu fie từ ram xuống bộ nhớ
            wb.save(self.file_bao_cao)
            messagebox.showinfo("Thành công", f"Đã xuất báo cáo ra file Excel tại: {self.file_bao_cao}")
        except Exception as e:
            messagebox.showerror("Lỗi", f"Không thể xuất báo cáo\n{e}\nVui lòng đóng file Excel trước khi xuất báo cáo")

    # Phương thức khởi tạo dữ liệu và gán chức năng vào các nút sau khi đăng nhập thành công
    def dang_nhap_thanh_cong(self):
        self.quan_ly_du_lieu.khoi_tao_du_lieu() # khởi tạo hoặc kết nối file
        self.cap_nhat_du_lieu() # Tải dữ liệu lên bảng hiển thị
        self.xu_ly_tim_kiem()
        self.nut_hien_thi.config(command=self.hien_thi_tat_ca)
        self.nut_hoan_thanh.config(command=self.hoan_thanh_don)
        self.nut_xoa_don.config(command=self.xoa_don_hang)
        self.nut_them_don.config(command=self.them_don_hang)
        self.nut_xuat_bao_cao.config(command=self.xuat_bao_cao)
        self.nut_sua_don.config(command=self.sua_don_hang)
      
    # Phương thức khở động chạy chương trình  
    def khoi_dong(self):
        Cua_So_Dang_Nhap(self)
        self.giao_dien.mainloop()
    
# Tạo đối tượng ứng dụng và chạy chương trình
ung_dung = Quan_Ly_Don_Hang()
ung_dung.khoi_dong()